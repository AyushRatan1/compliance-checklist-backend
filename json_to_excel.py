import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import re

def main():
    # Read the JSON file
    with open('final.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Convert to DataFrame
    df = pd.DataFrame(data)
    
    # Create Excel file with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Checklist"
    
    # Add headers
    headers = ['Checklist Name', 'Category', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for idx, item in enumerate(data, 2):  # Start from row 2
        ws.cell(row=idx, column=1, value=item['checklist_name'])
        ws.cell(row=idx, column=2, value=item['checklist_category'])
        ws.cell(row=idx, column=3, value=item['checklist_ai_description'])
        
        # Format description cell
        desc_cell = ws.cell(row=idx, column=3)
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50  # Checklist Name
    ws.column_dimensions['B'].width = 25  # Category
    ws.column_dimensions['C'].width = 100  # Description
    
    # Auto-adjust row heights for better readability
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = None  # Auto height
    
    # Save the Excel file
    filename = 'Compliance_Checklist.xlsx'
    wb.save(filename)
    print(f"Excel file '{filename}' has been created successfully!")
    print(f"Total checklist items: {len(data)}")
    
    # Also create a simple CSV version for compatibility
    df_simple = pd.DataFrame({
        'Checklist_Name': [item['checklist_name'] for item in data],
        'Category': [item['checklist_category'] for item in data],
        'Description': [item['checklist_ai_description'] for item in data]
    })
    
    csv_filename = 'Compliance_Checklist.csv'
    df_simple.to_csv(csv_filename, index=False, encoding='utf-8')
    print(f"CSV file '{csv_filename}' has also been created for compatibility!")

if __name__ == "__main__":
    main() 